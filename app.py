from flask import Flask, render_template, request, jsonify, send_file
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, date
import os
import io
import csv
import json
import secrets

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def format_date(val):
    """Format datetime as DD-MM-YYYY. Return empty string for None."""
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%d-%m-%Y')
    return str(val).strip()

def safe_str(val):
    """Return stripped string or empty string."""
    if val is None:
        return ''
    return str(val).strip()

def is_valid_id(val, skip_values=None):
    """Check if a cell value is a real employee ID (not a header or empty)."""
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    skip = skip_values or ['Emp ID', 'EMP ID', 'Employee ID', 'HKID', 'S NO', 'S N0', 'Sno']
    if s in skip:
        return False
    return True

def make_employee(emp_id, emp_name, doj, status, lwd, lop, remarks, source,
                  project='', centre='', designation=''):
    return {
        'emp_id':      emp_id,
        'emp_name':    safe_str(emp_name),
        'doj':         format_date(doj),
        'status':      status,
        'lwd':         format_date(lwd),
        'lop':         safe_str(lop),
        'remarks':     safe_str(remarks),
        'source':      source,
        'also_in':     '',
        'project':     safe_str(project),
        'centre':      safe_str(centre),
        'designation': safe_str(designation),
    }

def detect_cols(row):
    """Detect standard column positions from a header row.
    Accepts Cell objects (ws[N]) or plain values (iter_rows values_only=True).
    Returns dict — only found keys included; callers set their own fallbacks."""
    col = {}
    for idx, cell in enumerate(row):
        raw = cell.value if hasattr(cell, 'value') else cell
        h = safe_str(raw).upper()
        if not h:
            continue
        if 'id' not in col and any(x in h for x in ('EMP ID', 'EMPLOYEE ID', 'NEW EMP ID', 'HKID')):
            if 'OLD' not in h and 'PREV' not in h:
                col['id'] = idx
        if 'name' not in col and any(x in h for x in ('EMP NAME', 'EMPLOYEE NAME', 'HK STAFF NAME', 'STAFF NAME')):
            col['name'] = idx
        if 'doj' not in col and ('DATE OF JOINING' in h or h == 'DOJ'):
            col['doj'] = idx
        if 'lwd' not in col and ('LAST WORKING DAY' in h or h == 'LWD'):
            col['lwd'] = idx
        if 'lop' not in col and ('LOSS OF PAY' in h or 'LOP DAYS' in h or
                                  'LOP' in h or h == 'FINAL'):
            col['lop'] = idx
        if 'status' not in col and 'STATUS' in h:
            col['status'] = idx
        if 'rem' not in col and 'REMARKS' in h:
            col['rem'] = idx
        # New columns
        if 'proj' not in col and any(x in h for x in ('PROJECT', 'DEPARTMENT', 'DEPT')):
            col['proj'] = idx
        if 'centre' not in col and any(x in h for x in ('LOCATION', 'CENTRE', 'CENTER')):
            col['centre'] = idx
        if 'desig' not in col and any(x in h for x in ('DESIGNATION', 'DESIG')):
            col['desig'] = idx
    return col

# ---------------------------------------------------------------------------
# CORE
# ---------------------------------------------------------------------------
def find_tab(wb, patterns):
    """Find first tab matching any pattern (case-insensitive)."""
    for name in wb.sheetnames:
        name_upper = name.upper()
        for pat in patterns:
            if pat.upper() in name_upper:
                return name
    return None

def parse_core_file(filepath):
    """
    Active tab   : 'Core Staff <Month>- 2026'  — row 1=title, row 2=header, row 3+=data
    Secondary tab: 'Resigned & LOP'
      Section 1 header : 'RESIGNED EMPLOYEES DETAILS'
        cols (0-based): 0=SNo, 1=EmpID, 2=Name, 3=Center, 4=Desig, 5=Dept, 6=DOJ, 7=DOR, 8=LWD, 9=Remarks
      Section 2 header : 'LOP Details'  (or 'LOP' + 'DETAILS')
        cols: 0=SNo, 1=EmpID, 2=Name, 3=LOP_Days, 4=Location, 5=Remarks
    """
    employees = {}
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # --- Active tab ---
    active_tab = find_tab(wb, ['Core Staff', 'Core'])
    if not active_tab:
        raise ValueError("Core: Could not find 'Core Staff' tab")
    ws = wb[active_tab]
    col = detect_cols(ws[2])
    if 'id'   not in col: col['id']   = 1
    if 'name' not in col: col['name'] = 3
    if 'doj'  not in col: col['doj']  = 6
    if 'rem'  not in col: col['rem']  = 10

    for row in ws.iter_rows(min_row=3, values_only=True):
        raw_id = row[col['id']] if len(row) > col['id'] else None
        if not is_valid_id(raw_id):
            continue
        emp_id = safe_str(raw_id)
        employees[emp_id] = make_employee(
            emp_id      = emp_id,
            emp_name    = row[col['name']]  if len(row) > col['name']         else '',
            doj         = row[col['doj']]   if len(row) > col['doj']          else None,
            status      = 'Active',
            lwd         = None,
            lop         = '',
            remarks     = row[col['rem']]   if len(row) > col['rem']          else '',
            source      = 'Core',
            project     = row[col['proj']]  if 'proj'  in col and len(row) > col['proj']  else '',
            centre      = row[col['centre']]if 'centre'in col and len(row) > col['centre']else '',
            designation = row[col['desig']] if 'desig' in col and len(row) > col['desig'] else '',
        )

    # --- Resigned & LOP tab ---
    resigned_tab = find_tab(wb, ['Resigned', 'LOP'])
    if not resigned_tab:
        wb.close()
        return list(employees.values())
    ws = wb[resigned_tab]
    section = None
    header_skipped = False
    sec_col = {}
    for row in ws.iter_rows(values_only=True):
        if not any(c is not None for c in row):
            continue
        row_text = ' '.join(safe_str(c).upper() for c in row)

        if 'RESIGNED EMPLOYEES DETAILS' in row_text:
            section = 'resigned'; header_skipped = False; continue
        if ('LOP' in row_text and 'DETAIL' in row_text) or row_text.strip().startswith('LOP DETAILS'):
            section = 'lop'; header_skipped = False; continue

        if not header_skipped:
            header_skipped = True
            sec_col = detect_cols(row)
            continue

        if section == 'resigned':
            id_c   = sec_col.get('id', 1);  name_c = sec_col.get('name', 2)
            doj_c  = sec_col.get('doj', 6); lwd_c  = sec_col.get('lwd', 8)
            rem_c  = sec_col.get('rem', 9)
            emp_id = safe_str(row[id_c]) if len(row) > id_c else ''
            if not emp_id: continue
            emp_name  = row[name_c] if len(row) > name_c else ''
            doj       = row[doj_c]  if len(row) > doj_c  else None
            lwd       = row[lwd_c]  if len(row) > lwd_c  else None
            remarks   = row[rem_c]  if len(row) > rem_c  else ''
            proj_val  = safe_str(row[sec_col['proj']])   if 'proj'  in sec_col and len(row) > sec_col['proj']  else ''
            ctr_val   = safe_str(row[sec_col['centre']]) if 'centre'in sec_col and len(row) > sec_col['centre']else ''
            desig_val = safe_str(row[sec_col['desig']])  if 'desig' in sec_col and len(row) > sec_col['desig'] else ''
            if emp_id in employees:
                employees[emp_id]['status'] = 'Hold'
                employees[emp_id]['lwd']    = format_date(lwd)
                if remarks:   employees[emp_id]['remarks']     = safe_str(remarks)
                if proj_val  and not employees[emp_id].get('project'):     employees[emp_id]['project']     = proj_val
                if ctr_val   and not employees[emp_id].get('centre'):      employees[emp_id]['centre']      = ctr_val
                if desig_val and not employees[emp_id].get('designation'): employees[emp_id]['designation'] = desig_val
            else:
                employees[emp_id] = make_employee(emp_id, emp_name, doj, 'Hold', lwd, '', remarks, 'Core',
                                                  project=proj_val, centre=ctr_val, designation=desig_val)

        elif section == 'lop':
            id_c  = sec_col.get('id', 1)
            lop_c = sec_col.get('lop', 3)
            emp_id   = safe_str(row[id_c]) if len(row) > id_c else ''
            lop_days = row[lop_c] if len(row) > lop_c else ''
            if emp_id and emp_id in employees:
                employees[emp_id]['lop'] = safe_str(lop_days)

    wb.close()
    return list(employees.values())


# ---------------------------------------------------------------------------
# PROJECT
# ---------------------------------------------------------------------------
def parse_project_file(filepath):
    """
    Active tab   : 'Salary Request <Month> - 2026'  — row 1=title, row 2=header, row 3+=data
      cols (0-based): 0=SNo,1=EmpID,2=Slab,3=Name,4=Location,5=Dept,6=Desig,7=DOJ,8=Status
    Secondary tab: 'Resignations & LOP'
      Section 1 header: 'RESIGNATIONS'
        cols: 0=SNo,1=EmpID,2=Name,3=Location,4=Project,5=Desig,6=DOJ,7=DOR,8=LWD,9=Remarks
      Section 2 header: 'LOP Details'
        cols: 0=SNo,1=EmpID,2=Name,3=LOP_Days
    """
    employees = {}
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # --- Active tab ---
    active_tab = find_tab(wb, ['Salary Request', 'Project'])
    if not active_tab:
        raise ValueError("Project: Could not find 'Salary Request' or 'Project' tab")
    ws = wb[active_tab]
    col = detect_cols(ws[2])
    if 'id'     not in col: col['id']     = 1
    if 'name'   not in col: col['name']   = 3
    if 'doj'    not in col: col['doj']    = 7
    if 'status' not in col: col['status'] = 8

    for row in ws.iter_rows(min_row=3, values_only=True):
        raw_id = row[col['id']] if len(row) > col['id'] else None
        if not is_valid_id(raw_id):
            continue
        emp_id = safe_str(raw_id)
        employees[emp_id] = make_employee(
            emp_id      = emp_id,
            emp_name    = row[col['name']]    if len(row) > col['name']          else '',
            doj         = row[col['doj']]     if len(row) > col['doj']           else None,
            status      = safe_str(row[col['status']]) if len(row) > col['status'] else 'Active',
            lwd         = None,
            lop         = '',
            remarks     = '',
            source      = 'Project',
            project     = row[col['proj']]   if 'proj'  in col and len(row) > col['proj']  else '',
            centre      = row[col['centre']] if 'centre'in col and len(row) > col['centre']else '',
            designation = row[col['desig']]  if 'desig' in col and len(row) > col['desig'] else '',
        )

    # --- Resignations & LOP tab ---
    resigned_tab = find_tab(wb, ['Resignations', 'LOP', 'Left'])
    if not resigned_tab:
        wb.close()
        return list(employees.values())
    ws = wb[resigned_tab]
    section = None
    header_skipped = False
    sec_col = {}
    for row in ws.iter_rows(values_only=True):
        if not any(c is not None for c in row):
            continue
        row_text = ' '.join(safe_str(c).upper() for c in row)

        if 'RESIGNATIONS' in row_text and section is None:
            section = 'resigned'; header_skipped = False; continue
        if 'LOP' in row_text and 'DETAIL' in row_text:
            section = 'lop'; header_skipped = False; continue

        if not header_skipped:
            header_skipped = True
            sec_col = detect_cols(row)
            continue

        if section == 'resigned':
            id_c   = sec_col.get('id', 1);  name_c = sec_col.get('name', 2)
            doj_c  = sec_col.get('doj', 6); lwd_c  = sec_col.get('lwd', 8)
            rem_c  = sec_col.get('rem', 9)
            emp_id = safe_str(row[id_c]) if len(row) > id_c else ''
            if not emp_id: continue
            emp_name  = row[name_c] if len(row) > name_c else ''
            doj       = row[doj_c]  if len(row) > doj_c  else None
            lwd       = row[lwd_c]  if len(row) > lwd_c  else None
            remarks   = row[rem_c]  if len(row) > rem_c  else ''
            proj_val  = safe_str(row[sec_col['proj']])   if 'proj'  in sec_col and len(row) > sec_col['proj']  else ''
            ctr_val   = safe_str(row[sec_col['centre']]) if 'centre'in sec_col and len(row) > sec_col['centre']else ''
            desig_val = safe_str(row[sec_col['desig']])  if 'desig' in sec_col and len(row) > sec_col['desig'] else ''
            if emp_id in employees:
                employees[emp_id]['status'] = 'Hold'
                employees[emp_id]['lwd']    = format_date(lwd)
                if proj_val  and not employees[emp_id].get('project'):     employees[emp_id]['project']     = proj_val
                if ctr_val   and not employees[emp_id].get('centre'):      employees[emp_id]['centre']      = ctr_val
                if desig_val and not employees[emp_id].get('designation'): employees[emp_id]['designation'] = desig_val
            else:
                employees[emp_id] = make_employee(emp_id, emp_name, doj, 'Hold', lwd, '', remarks, 'Project',
                                                  project=proj_val, centre=ctr_val, designation=desig_val)

        elif section == 'lop':
            id_c  = sec_col.get('id', 1)
            lop_c = sec_col.get('lop', 3)
            emp_id   = safe_str(row[id_c]) if len(row) > id_c else ''
            lop_days = row[lop_c] if len(row) > lop_c else ''
            if emp_id and emp_id in employees:
                employees[emp_id]['lop'] = safe_str(lop_days)

    wb.close()
    return list(employees.values())


# ---------------------------------------------------------------------------
# CF
# ---------------------------------------------------------------------------
def parse_cf_file(filepath):
    """
    Active tab   : 'Active <Month> 2026' — row 1=title, row 2=header, row 3+=data
      cols: 0=SNo,1=EmpID,2=Name,3=Location,4=Desig,5=DOJ,6=Dept,7=Territory,8=Region,9=Status,10=Remarks
    Secondary tab: 'Resigned'
      Section 1 header: 'RESIGNATIONS'
        cols: 0=SNo,1=EmpID,2=Name,3=Location,4=Project,5=Desig,6=DOJ,7=DOR,8=LWD,9=Remarks
      Section 2 header: 'LOP'
        cols: 0=SNo,1=EmpID,2=Name,3=LOP_Days,4=Location,5=Remarks
    """
    employees = {}
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # --- Active tab ---
    active_tab = find_tab(wb, ['Active'])
    if not active_tab:
        raise ValueError("CF: Could not find 'Active' tab")
    ws = wb[active_tab]
    col = detect_cols(ws[2])
    if 'id'   not in col: col['id']   = 1
    if 'name' not in col: col['name'] = 2
    if 'doj'  not in col: col['doj']  = 5
    if 'rem'  not in col: col['rem']  = 10

    for row in ws.iter_rows(min_row=3, values_only=True):
        raw_id = row[col['id']] if len(row) > col['id'] else None
        if not is_valid_id(raw_id):
            continue
        emp_id = safe_str(raw_id)
        employees[emp_id] = make_employee(
            emp_id      = emp_id,
            emp_name    = row[col['name']]   if len(row) > col['name']          else '',
            doj         = row[col['doj']]    if len(row) > col['doj']           else None,
            status      = 'Active',
            lwd         = None,
            lop         = '',
            remarks     = row[col['rem']]    if len(row) > col['rem']           else '',
            source      = 'CF',
            project     = row[col['proj']]   if 'proj'  in col and len(row) > col['proj']  else '',
            centre      = row[col['centre']] if 'centre'in col and len(row) > col['centre']else '',
            designation = row[col['desig']]  if 'desig' in col and len(row) > col['desig'] else '',
        )

    # --- Resigned tab ---
    resigned_tab = find_tab(wb, ['Resigned', 'Left'])
    if not resigned_tab:
        wb.close()
        return list(employees.values())
    ws = wb[resigned_tab]
    section = None
    header_skipped = False
    sec_col = {}
    for row in ws.iter_rows(values_only=True):
        if not any(c is not None for c in row):
            continue
        row_text = ' '.join(safe_str(c).upper() for c in row)

        if 'RESIGNATIONS' in row_text and section is None:
            section = 'resigned'; header_skipped = False; continue
        # CF uses single 'LOP' keyword — also catch 'LOP Details' format
        non_null = [safe_str(c).upper() for c in row if c is not None]
        if non_null == ['LOP'] or ('LOP' in row_text and 'DETAIL' in row_text):
            section = 'lop'; header_skipped = False; continue

        if not header_skipped:
            header_skipped = True
            sec_col = detect_cols(row)
            continue

        if section == 'resigned':
            id_c   = sec_col.get('id', 1);  name_c = sec_col.get('name', 2)
            doj_c  = sec_col.get('doj', 6); lwd_c  = sec_col.get('lwd', 8)
            rem_c  = sec_col.get('rem', 9)
            emp_id = safe_str(row[id_c]) if len(row) > id_c else ''
            if not emp_id: continue
            emp_name  = row[name_c] if len(row) > name_c else ''
            doj       = row[doj_c]  if len(row) > doj_c  else None
            lwd       = row[lwd_c]  if len(row) > lwd_c  else None
            remarks   = row[rem_c]  if len(row) > rem_c  else ''
            proj_val  = safe_str(row[sec_col['proj']])   if 'proj'  in sec_col and len(row) > sec_col['proj']  else ''
            ctr_val   = safe_str(row[sec_col['centre']]) if 'centre'in sec_col and len(row) > sec_col['centre']else ''
            desig_val = safe_str(row[sec_col['desig']])  if 'desig' in sec_col and len(row) > sec_col['desig'] else ''
            if emp_id in employees:
                employees[emp_id]['status'] = 'Hold'
                employees[emp_id]['lwd']    = format_date(lwd)
                if proj_val  and not employees[emp_id].get('project'):     employees[emp_id]['project']     = proj_val
                if ctr_val   and not employees[emp_id].get('centre'):      employees[emp_id]['centre']      = ctr_val
                if desig_val and not employees[emp_id].get('designation'): employees[emp_id]['designation'] = desig_val
            else:
                employees[emp_id] = make_employee(emp_id, emp_name, doj, 'Hold', lwd, '', remarks, 'CF',
                                                  project=proj_val, centre=ctr_val, designation=desig_val)

        elif section == 'lop':
            id_c  = sec_col.get('id', 1)
            lop_c = sec_col.get('lop', 3)
            emp_id   = safe_str(row[id_c]) if len(row) > id_c else ''
            lop_days = row[lop_c] if len(row) > lop_c else ''
            if emp_id and emp_id in employees:
                employees[emp_id]['lop'] = safe_str(lop_days)

    wb.close()
    return list(employees.values())


# ---------------------------------------------------------------------------
# HK
# ---------------------------------------------------------------------------
def parse_hk_file(filepath):
    """
    Active tab : '<Month>-26' or '<Month>-2026' — row 1=title, row 2=header, row 3+=data
      Status values: 'Working' = Active, 'New' = Active, 'Left' = Hold
      LOP comes from 'Loss Of Pay' column directly in active tab (col 45, 0-based=44)
      LWD comes from 'Last Working Day' column (col 8, 0-based=7)
      Remarks col 48 (0-based=47)
    Left tab: 'Left' — row 1=title, row 2=header, row 3+=data → status=Hold
    SKIP: 'LOP' tab, 'Left F&F' tab, 'New HK staff' tab
    """
    employees = {}
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # --- Active tab ---
    active_tab = find_tab(wb, ['HK', 'Mar', 'Feb', 'Jan', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'])
    if not active_tab:
        # Try first tab that looks like a month tab
        for name in wb.sheetnames:
            if any(x in name for x in ['-26', '-2026', '2026', 'Staff', 'Attendance']):
                active_tab = name
                break
    if not active_tab:
        raise ValueError("HK: Could not find month data tab")
    ws = wb[active_tab]
    # Detect columns from header row 2
    col = {}
    for idx, cell in enumerate(ws[2]):
        h = safe_str(cell.value).upper()
        if h == 'HKID':                                col['id']     = idx
        elif 'HK STAFF NAME' in h or 'STAFF NAME' in h: col['name']  = idx
        elif 'DATE OF JOINING' in h:                   col['doj']    = idx
        elif 'LAST WORKING DAY' in h:                  col['lwd']    = idx
        elif 'LOSS OF PAY' in h:                       col['lop']    = idx
        elif 'REMARKS' in h:                           col['rem']    = idx
        elif 'EMP STATUS' in h:                        col['status'] = idx

    for row in ws.iter_rows(min_row=3, values_only=True):
        raw_id = row[col['id']] if 'id' in col else None
        if not is_valid_id(raw_id):
            continue
        emp_id = safe_str(raw_id)
        status_raw = safe_str(row[col['status']] if 'status' in col else '').upper()
        # 'Working' and 'New' are both Active employees
        status = 'Active' if status_raw in ('WORKING', 'NEW', 'ACTIVE') else 'Hold'

        employees[emp_id] = make_employee(
            emp_id   = emp_id,
            emp_name = row[col['name']] if 'name' in col else '',
            doj      = row[col['doj']]  if 'doj'  in col else None,
            status   = status,
            lwd      = row[col['lwd']]  if 'lwd'  in col else None,
            lop      = row[col['lop']]  if 'lop'  in col else '',
            remarks  = row[col['rem']]  if 'rem'  in col else '',
            source   = 'HK',
        )

    # --- Left tab — employees who left mid-month ---
    left_tab = find_tab(wb, ['Left', 'Resigned'])
    if not left_tab:
        wb.close()
        return list(employees.values())
    ws = wb[left_tab]
    # Same column structure as active tab, detect from row 2
    col_left = {}
    for idx, cell in enumerate(ws[2]):
        h = safe_str(cell.value).upper()
        if h == 'HKID':                                  col_left['id']   = idx
        elif 'HK STAFF NAME' in h or 'STAFF NAME' in h: col_left['name'] = idx
        elif 'DATE OF JOINING' in h:                     col_left['doj']  = idx
        elif 'LAST WORKING DAY' in h:                    col_left['lwd']  = idx

    for row in ws.iter_rows(min_row=3, values_only=True):
        raw_id = row[col_left['id']] if 'id' in col_left else None
        if not is_valid_id(raw_id):
            continue
        emp_id = safe_str(raw_id)
        lwd    = row[col_left['lwd']] if 'lwd' in col_left else None
        if emp_id in employees:
            employees[emp_id]['status'] = 'Hold'
            if lwd: employees[emp_id]['lwd'] = format_date(lwd)
        else:
            employees[emp_id] = make_employee(
                emp_id   = emp_id,
                emp_name = row[col_left['name']] if 'name' in col_left else '',
                doj      = row[col_left['doj']]  if 'doj'  in col_left else None,
                status   = 'Hold',
                lwd      = lwd,
                lop      = '',
                remarks  = '',
                source   = 'HK',
            )

    # SKIP: 'LOP' tab (primary LOP already in active tab Loss Of Pay column)
    # SKIP: 'Left F&F' tab (finance only, same employees as Left tab)
    # SKIP: 'New HK staff' tab (already in Mar-26 active tab)

    wb.close()
    return list(employees.values())


# ---------------------------------------------------------------------------
# RETAINER
# ---------------------------------------------------------------------------
def parse_retainer_file(filepath):
    """
    Active tab  : '<Month> 26' — row 1=header (NO title row), row 2+=data
      cols: 0=SNo,1=Month,2=EmpID,3=Name,4=Location,5=Project,6=Desig,7=DOJ,8=LWD,9=Status,10=Remarks
    LOP tab     : 'LOP' — row 1=header (NO title row), row 2+=data
      cols: 0=EmpID,1=Name,2=Location,3=LOP_Days(Final)
    Hold tab    : 'Hold salary - F&F to be shared' — row 1=header (NO title row), row 2+=data
      cols: 0=SNo,1=Month,2=EmpID,3=Name,4=Location,5=Project,6=Desig,7=DOJ,8=LWD,9=Remarks
    """
    employees = {}
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # --- Active tab (NO title row — header is row 1, data from row 2) ---
    active_tab = find_tab(wb, ['Retainer', '26', 'Active'])
    if not active_tab:
        # Try to find a tab with month-like name
        for name in wb.sheetnames:
            if any(x in name.lower() for x in ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec', '26']):
                active_tab = name
                break
    if not active_tab:
        raise ValueError("Retainer: Could not find month data tab")
    ws = wb[active_tab]
    col = detect_cols(ws[1])  # no title row — header is row 1
    if 'id'     not in col: col['id']     = 2
    if 'name'   not in col: col['name']   = 3
    if 'doj'    not in col: col['doj']    = 7
    if 'lwd'    not in col: col['lwd']    = 8
    if 'status' not in col: col['status'] = 9
    if 'rem'    not in col: col['rem']    = 10

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_id = row[col['id']] if len(row) > col['id'] else None
        if not is_valid_id(raw_id):
            continue
        emp_id = safe_str(raw_id)
        employees[emp_id] = make_employee(
            emp_id      = emp_id,
            emp_name    = row[col['name']]    if len(row) > col['name']          else '',
            doj         = row[col['doj']]     if len(row) > col['doj']           else None,
            status      = safe_str(row[col['status']]) if len(row) > col['status'] else 'Active',
            lwd         = row[col['lwd']]     if len(row) > col['lwd']           else None,
            lop         = '',
            remarks     = row[col['rem']]     if len(row) > col['rem']           else '',
            source      = 'Retainer',
            project     = row[col['proj']]   if 'proj'  in col and len(row) > col['proj']  else '',
            centre      = row[col['centre']] if 'centre'in col and len(row) > col['centre']else '',
            designation = row[col['desig']]  if 'desig' in col and len(row) > col['desig'] else '',
        )

    # --- LOP tab (NO title row — header is row 1, data from row 2) ---
    lop_tab = find_tab(wb, ['LOP'])
    if lop_tab:
        ws = wb[lop_tab]
        lop_col = detect_cols(ws[1])
        id_c  = lop_col.get('id',  0)
        lop_c = lop_col.get('lop', 3)
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_id = row[id_c] if len(row) > id_c else None
            if not is_valid_id(raw_id):
                continue
            emp_id   = safe_str(raw_id)
            lop_days = row[lop_c] if len(row) > lop_c else ''
            if emp_id in employees:
                employees[emp_id]['lop'] = safe_str(lop_days)

    # --- Hold tab (NO title row — header is row 1, data from row 2) ---
    hold_tab = find_tab(wb, ['Hold', 'F&F'])
    if hold_tab:
        ws = wb[hold_tab]
        hcol = detect_cols(ws[1])
        if 'id'   not in hcol: hcol['id']   = 2
        if 'name' not in hcol: hcol['name'] = 3
        if 'doj'  not in hcol: hcol['doj']  = 7
        if 'lwd'  not in hcol: hcol['lwd']  = 8
        if 'rem'  not in hcol: hcol['rem']  = 9
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_id = row[hcol['id']] if len(row) > hcol['id'] else None
            if not is_valid_id(raw_id):
                continue
            emp_id    = safe_str(raw_id)
            emp_name  = row[hcol['name']] if len(row) > hcol['name'] else ''
            doj       = row[hcol['doj']]  if len(row) > hcol['doj']  else None
            lwd       = row[hcol['lwd']]  if len(row) > hcol['lwd']  else None
            remarks   = row[hcol['rem']]  if len(row) > hcol['rem']  else ''
            proj_val  = safe_str(row[hcol['proj']])   if 'proj'  in hcol and len(row) > hcol['proj']  else ''
            ctr_val   = safe_str(row[hcol['centre']]) if 'centre'in hcol and len(row) > hcol['centre']else ''
            desig_val = safe_str(row[hcol['desig']])  if 'desig' in hcol and len(row) > hcol['desig'] else ''
            if emp_id in employees:
                employees[emp_id]['status'] = 'Hold'
                if lwd: employees[emp_id]['lwd'] = format_date(lwd)
                if proj_val  and not employees[emp_id].get('project'):     employees[emp_id]['project']     = proj_val
                if ctr_val   and not employees[emp_id].get('centre'):      employees[emp_id]['centre']      = ctr_val
                if desig_val and not employees[emp_id].get('designation'): employees[emp_id]['designation'] = desig_val
            else:
                employees[emp_id] = make_employee(emp_id, emp_name, doj, 'Hold', lwd, '', remarks, 'Retainer',
                                                  project=proj_val, centre=ctr_val, designation=desig_val)

    wb.close()
    return list(employees.values())


# ---------------------------------------------------------------------------
# SCHOOL
# ---------------------------------------------------------------------------
def parse_school_file(filepath):
    """
    Active tab   : 'KARV Salary <Month> - 2026' — row 1=title, row 2=header, row 3+=data
      cols: 0=SNo,1=OldID,2=EmpID(KARV/KARVJC/DRF),3=Name,4=Program,5=Dept,6=Desig,7=DOJ,8=Status,9=Remarks
    Secondary tab: 'Resignations & LOP'
      Section 1 header: 'RESIGNATIONS'
        cols: 0=SNo,1=EmpID,2=Name,3=Program,4=Dept,5=Desig,6=DOJ,7=DOR,8=LWD,9=Remarks
      Section 2 header: 'LOP Details'
        cols: 0=SNo,1=EmpID,2=Name,3=LOP_Days
    """
    employees = {}
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # --- Active tab ---
    active_tab = find_tab(wb, ['KARV', 'School', 'Salary'])
    if not active_tab:
        raise ValueError("School: Could not find 'KARV Salary' or similar tab")
    ws = wb[active_tab]
    col = detect_cols(ws[2])
    if 'id'     not in col: col['id']     = 2  # col 0=SNo, col 1=Old ID, col 2=Emp ID
    if 'name'   not in col: col['name']   = 3
    if 'doj'    not in col: col['doj']    = 7
    if 'status' not in col: col['status'] = 8
    if 'rem'    not in col: col['rem']    = 9

    for row in ws.iter_rows(min_row=3, values_only=True):
        raw_id = row[col['id']] if len(row) > col['id'] else None
        if not is_valid_id(raw_id):
            continue
        emp_id = safe_str(raw_id)
        employees[emp_id] = make_employee(
            emp_id   = emp_id,
            emp_name = row[col['name']]   if len(row) > col['name']   else '',
            doj      = row[col['doj']]    if len(row) > col['doj']    else None,
            status   = safe_str(row[col['status']]) if len(row) > col['status'] else 'Active',
            lwd      = None,
            lop      = '',
            remarks  = row[col['rem']]    if len(row) > col['rem']    else '',
            source   = 'School',
        )

    # --- Resignations & LOP tab ---
    resigned_tab = find_tab(wb, ['Resignations', 'LOP', 'Left'])
    if not resigned_tab:
        wb.close()
        return list(employees.values())
    ws = wb[resigned_tab]
    section = None
    header_skipped = False
    sec_col = {}
    for row in ws.iter_rows(values_only=True):
        if not any(c is not None for c in row):
            continue
        row_text = ' '.join(safe_str(c).upper() for c in row)

        if 'RESIGNATIONS' in row_text and section is None:
            section = 'resigned'; header_skipped = False; continue
        if 'LOP' in row_text and 'DETAIL' in row_text:
            section = 'lop'; header_skipped = False; continue

        if not header_skipped:
            header_skipped = True
            sec_col = detect_cols(row)
            continue

        if section == 'resigned':
            id_c   = sec_col.get('id', 1);  name_c = sec_col.get('name', 2)
            doj_c  = sec_col.get('doj', 6); lwd_c  = sec_col.get('lwd', 8)
            rem_c  = sec_col.get('rem', 9)
            emp_id = safe_str(row[id_c]) if len(row) > id_c else ''
            if not emp_id: continue
            emp_name = row[name_c] if len(row) > name_c else ''
            doj      = row[doj_c]  if len(row) > doj_c  else None
            lwd      = row[lwd_c]  if len(row) > lwd_c  else None
            remarks  = row[rem_c]  if len(row) > rem_c  else ''
            if emp_id in employees:
                employees[emp_id]['status'] = 'Hold'
                employees[emp_id]['lwd']    = format_date(lwd)
            else:
                employees[emp_id] = make_employee(emp_id, emp_name, doj, 'Hold', lwd, '', remarks, 'School')

        elif section == 'lop':
            id_c  = sec_col.get('id', 1)
            lop_c = sec_col.get('lop', 3)
            emp_id   = safe_str(row[id_c]) if len(row) > id_c else ''
            lop_days = row[lop_c] if len(row) > lop_c else ''
            if emp_id and emp_id in employees:
                employees[emp_id]['lop'] = safe_str(lop_days)

    wb.close()
    return list(employees.values())


# ---------------------------------------------------------------------------
# COLLEGE
# ---------------------------------------------------------------------------
def parse_college_file(filepath):
    """
    Active tab (ONLY tab): 'Active' — row 1=header (NO title row), row 2+=data
      cols: 0=SNo,1=Month,2=EmpID,3=Name,4=Project,5=Desig,6=DOJ,7=LWD,8=LOP_Days,9=Status,
            10=Consolidated_Fees,11=Other_Allowance,12=Remarks
    """
    employees = {}
    wb = openpyxl.load_workbook(filepath, data_only=True)

    active_tab = find_tab(wb, ['Active'])
    if not active_tab:
        raise ValueError("College: Could not find 'Active' tab")
    ws = wb[active_tab]
    col = detect_cols(ws[1])  # no title row — header is row 1
    if 'id'     not in col: col['id']     = 2
    if 'name'   not in col: col['name']   = 3
    if 'doj'    not in col: col['doj']    = 6
    if 'lwd'    not in col: col['lwd']    = 7
    if 'lop'    not in col: col['lop']    = 8
    if 'status' not in col: col['status'] = 9
    if 'rem'    not in col: col['rem']    = 12

    for row in ws.iter_rows(min_row=2, values_only=True):  # NO title row → data from row 2
        raw_id = row[col['id']] if len(row) > col['id'] else None
        if not is_valid_id(raw_id):
            continue
        emp_id = safe_str(raw_id)
        employees[emp_id] = make_employee(
            emp_id   = emp_id,
            emp_name = row[col['name']]   if len(row) > col['name']   else '',
            doj      = row[col['doj']]    if len(row) > col['doj']    else None,
            status   = safe_str(row[col['status']]) if len(row) > col['status'] else 'Active',
            lwd      = row[col['lwd']]    if len(row) > col['lwd']    else None,
            lop      = row[col['lop']]    if len(row) > col['lop']    else '',
            remarks  = row[col['rem']]    if len(row) > col['rem']    else '',
            source   = 'College',
        )

    wb.close()
    return list(employees.values())


# ---------------------------------------------------------------------------
# STANDALONE LOP REPORT
# ---------------------------------------------------------------------------
def parse_lop_report_file(filepath):
    """
    Standalone LOP report with fixed structure.
    Row 1 = title ('LOP'), row 2 = header, row 3+ = data.
    Expected cols: S No | New Emp ID | Emp Name | LOP Days | Location | Category | Remarks
    Returns list of (emp_id, lop_days) tuples where lop_days > 0.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    result = []
    col = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # title row
        if i == 1:  # header row
            col = detect_cols(row)
            if 'id'  not in col: col['id']  = 1
            if 'lop' not in col: col['lop'] = 3
            continue
        raw_id = row[col['id']] if len(row) > col['id'] else None
        if not is_valid_id(raw_id):
            continue
        emp_id  = safe_str(raw_id)
        lop_raw = row[col['lop']] if len(row) > col['lop'] else None
        try:
            lop_val = float(str(lop_raw).strip()) if lop_raw is not None else 0
        except Exception:
            lop_val = 0
        if lop_val > 0:
            result.append((emp_id, int(lop_val)))
    wb.close()
    return result


# ---------------------------------------------------------------------------
# FLASK ROUTES
# ---------------------------------------------------------------------------
@app.route('/')
def index():
    return render_template('index.html')


@app.route('/process', methods=['POST'])
def process_files():
    files = {
        'core':     request.files.get('core'),
        'project':  request.files.get('project'),
        'cf':       request.files.get('cf'),
        'hk':       request.files.get('hk'),
        'retainer': request.files.get('retainer'),
        'school':   request.files.get('school'),
        'college':  request.files.get('college'),
    }
    lop_files = {k: request.files.get(f'{k}_lop')
                 for k in ['core','project','cf','hk','retainer','school','college']}

    parsers = {
        'core':     parse_core_file,
        'project':  parse_project_file,
        'cf':       parse_cf_file,
        'hk':       parse_hk_file,
        'retainer': parse_retainer_file,
        'school':   parse_school_file,
        'college':  parse_college_file,
    }

    all_employees = {}
    errors = []

    # Wipe all stale upload files from any previous session before saving new ones
    for fname in os.listdir(UPLOAD_FOLDER):
        if fname.startswith('upload_') and fname.endswith('.xlsx'):
            try:
                os.remove(os.path.join(UPLOAD_FOLDER, fname))
            except Exception:
                pass

    # Reset manifest to only track files uploaded in THIS request
    manifest_path = os.path.join(UPLOAD_FOLDER, 'manifest.json')
    uploaded_keys = [k for k, f in files.items() if f and allowed_file(f.filename)]
    for k, lf in lop_files.items():
        if lf and allowed_file(lf.filename):
            uploaded_keys.append(f'{k}_lop')
    with open(manifest_path, 'w') as mf:
        json.dump({'uploaded': uploaded_keys}, mf)

    for key, file in files.items():
        if not file or not allowed_file(file.filename):
            continue
        try:
            # Save with fixed key-based name so LOP route can find them reliably
            filepath = os.path.join(UPLOAD_FOLDER, f'upload_{key}.xlsx')
            file.save(filepath)
            employees = parsers[key](filepath)

            for emp in employees:
                eid = emp['emp_id']
                if eid in all_employees:
                    existing_source  = all_employees[eid]['source']
                    duplicate_source = emp['source']
                    if existing_source != duplicate_source:
                        already = all_employees[eid].get('also_in', '')
                        if already:
                            if duplicate_source not in already:
                                all_employees[eid]['also_in'] = already + ', ' + duplicate_source
                        else:
                            all_employees[eid]['also_in'] = duplicate_source
                    if emp.get('lop') and not all_employees[eid].get('lop'):
                        all_employees[eid]['lop'] = emp['lop']
                    if emp.get('lwd') and not all_employees[eid].get('lwd'):
                        all_employees[eid]['lwd'] = emp['lwd']
                    if emp['status'] == 'Hold':
                        all_employees[eid]['status'] = 'Hold'
                    # Fill in new fields if missing from primary source
                    for f in ('project', 'centre', 'designation'):
                        if emp.get(f) and not all_employees[eid].get(f):
                            all_employees[eid][f] = emp[f]
                else:
                    all_employees[eid] = emp

            # DO NOT delete files — kept for LOP CSV generation (Use Case 2)

        except Exception as e:
            errors.append(f'Error processing {key}: {str(e)}')

    # Apply per-sheet standalone LOP reports — override LOP values from salary sheets
    for key, lf in lop_files.items():
        if not lf or not allowed_file(lf.filename):
            continue
        try:
            fp = os.path.join(UPLOAD_FOLDER, f'upload_{key}_lop.xlsx')
            lf.save(fp)
            lop_records = parse_lop_report_file(fp)
            for emp_id, lop_days in lop_records:
                if emp_id in all_employees:
                    all_employees[emp_id]['lop'] = str(lop_days)
        except Exception as e:
            errors.append(f'Error processing {key.title()} LOP: {str(e)}')

    employees_list = list(all_employees.values())
    total        = len(employees_list)
    active_count = sum(1 for e in employees_list if e['status'] == 'Active')
    hold_count   = sum(1 for e in employees_list if e['status'] == 'Hold')

    return jsonify({
        'success':   len(errors) == 0,
        'employees': employees_list,
        'stats': {
            'total':  total,
            'active': active_count,
            'hold':   hold_count,
        },
        'errors': errors,
    })


@app.route('/download', methods=['POST'])
def download_excel():
    data      = request.get_json()
    employees = data.get('employees', [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Consolidated Data'

    headers = ['Emp ID', 'Emp Name', 'Date of Joining', 'Status', 'LOP',
               'Last Working Day', 'Project / Dept', 'Centre / Location',
               'Designation', 'Remarks', 'Source Sheet', 'Also In']
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    for emp in employees:
        ws.append([
            emp.get('emp_id',      ''),
            emp.get('emp_name',    ''),
            emp.get('doj',         ''),
            emp.get('status',      ''),
            emp.get('lop',         ''),
            emp.get('lwd',         ''),
            emp.get('project',     ''),
            emp.get('centre',      ''),
            emp.get('designation', ''),
            emp.get('remarks',     ''),
            emp.get('source',      ''),
            emp.get('also_in',     ''),
        ])

    for column in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in column), default=0)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(max_len + 2, 50)

    month_label  = datetime.now().strftime('%B_%Y')
    output_path  = os.path.join(UPLOAD_FOLDER, f'Employee_Consolidated_{month_label}.xlsx')
    wb.save(output_path)
    wb.close()

    return send_file(
        output_path,
        as_attachment=True,
        download_name=f'Employee_Consolidated_{month_label}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ---------------------------------------------------------------------------
# LOP HELPER — shared section parser
# ---------------------------------------------------------------------------
def parse_lop_section(ws, section1_keywords, section2_keywords):
    """
    Generic parser for tabs with two stacked sections (resigned + LOP).
    section1_keywords / section2_keywords can be a string or list of strings —
    any partial match (case-insensitive) triggers the section.
    Returns list of (emp_id, lop_days) tuples from the LOP section only.
    """
    # Normalise to lists
    if isinstance(section1_keywords, str): section1_keywords = [section1_keywords]
    if isinstance(section2_keywords, str): section2_keywords = [section2_keywords]
    s1 = [k.upper() for k in section1_keywords]
    s2 = [k.upper() for k in section2_keywords]

    result = []
    section = None
    header_skipped = False
    sec_col = {}
    for row in ws.iter_rows(values_only=True):
        if not any(c is not None for c in row):
            continue
        row_text = ' '.join(safe_str(c).upper() for c in row)
        if section is None and any(k in row_text for k in s1):
            section = 'first'; header_skipped = False; continue
        if any(k in row_text for k in s2):
            section = 'lop'; header_skipped = False; continue
        if not header_skipped:
            header_skipped = True
            sec_col = detect_cols(row)
            continue
        if section == 'lop':
            id_c  = sec_col.get('id', 1)
            lop_c = sec_col.get('lop', 3)
            emp_id = safe_str(row[id_c]) if len(row) > id_c else ''
            if not emp_id: continue
            lop_raw = row[lop_c] if len(row) > lop_c else None
            try:
                lop_val = float(str(lop_raw).strip()) if lop_raw is not None else 0
            except:
                lop_val = 0
            if lop_val > 0:
                result.append((emp_id, int(lop_val)))
    return result


def generate_lop_csv_content(records, lop_period, payout_period):
    """Generate CSV string from list of (emp_id, lop_days) tuples."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Email/Employee ID', 'LOP Period', 'Payout Period',
                     'LOP(Days)', 'LOP Reversal(Days)', 'Add/Delete'])
    for emp_id, lop_days in records:
        writer.writerow([emp_id, lop_period, payout_period, lop_days, '', 'Add'])
    return output.getvalue()


# ---------------------------------------------------------------------------
# LOP DOWNLOAD ROUTE
# ---------------------------------------------------------------------------
@app.route('/download_lop_csv', methods=['POST'])
def download_lop_csv():
    """
    Download a single LOP CSV.
    Expects JSON: { "type": "onroll" | "offroll", "lop_period": "2026-03", "payout_period": "2026-03" }
    """
    data         = request.get_json()
    csv_type     = data.get('type', '')          # 'onroll' or 'offroll'
    lop_period   = data.get('lop_period', '')
    payout_period = data.get('payout_period', '')

    if not lop_period or not payout_period:
        return jsonify({'error': 'LOP Period and Payout Period are required'}), 400

    # Use same parsers as the consolidated view — guarantees consistent LOP values
    SOURCES = {
        'onroll':  [('core', parse_core_file), ('project', parse_project_file),
                    ('hk',   parse_hk_file),   ('cf',      parse_cf_file)],
        'offroll': [('retainer', parse_retainer_file), ('school',  parse_school_file),
                    ('college',  parse_college_file)],
    }
    if csv_type not in SOURCES:
        return jsonify({'error': 'Invalid type — must be onroll or offroll'}), 400

    records = []
    errors  = []
    seen    = set()

    for key, parser in SOURCES[csv_type]:
        fp = os.path.join(UPLOAD_FOLDER, f'upload_{key}.xlsx')
        if not os.path.exists(fp):
            continue
        try:
            for emp in parser(fp):
                emp_id = emp['emp_id']
                if emp_id in seen:
                    continue
                try:
                    lop_val = float(str(emp.get('lop') or 0))
                except (ValueError, TypeError):
                    lop_val = 0
                if lop_val > 0:
                    records.append((emp_id, int(lop_val)))
                    seen.add(emp_id)
        except Exception as e:
            errors.append(f'{key.title()}: {e}')

    # Apply any per-sheet standalone LOP override files
    for key, _ in SOURCES[csv_type]:
        fp = os.path.join(UPLOAD_FOLDER, f'upload_{key}_lop.xlsx')
        if not os.path.exists(fp):
            continue
        try:
            for emp_id, lop_days in parse_lop_report_file(fp):
                records = [(e, d) for e, d in records if e != emp_id]
                records.append((emp_id, lop_days))
        except Exception as e:
            errors.append(f'{key.title()} LOP override: {e}')

    if errors:
        return jsonify({'error': 'Could not generate CSV due to errors: ' + '; '.join(errors)}), 400

    filename = (f'LOP_Onroll_{payout_period}.csv' if csv_type == 'onroll'
                else f'LOP_Offroll_{payout_period}.csv')
    csv_content = generate_lop_csv_content(records, lop_period, payout_period)
    output = io.BytesIO(csv_content.encode('utf-8'))
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='text/csv',
    )


# ---------------------------------------------------------------------------
# NEW JOINEE EXCEL EXPORT (Use Case 3)
# ---------------------------------------------------------------------------

def calc_amount(doj):
    """
    Financial year April–March: remaining months in FY × 20.
    April joiner → 240, May → 220, June → 200, ... March → 20.
    """
    if not isinstance(doj, (datetime, date)):
        return 20
    return (12 - ((doj.month - 4) % 12)) * 20


def get_new_joinees_from_file(filepath, tab_patterns, skip_rows, id_col_fallback, doj_col_fallback):
    """
    Detect new joinees using header-name based column detection.
    Falls back to positional indices if header not found.
    For unlabelled Remarks columns (like Project), scans ALL columns for 'New Joinee'.
    Returns list of (emp_id, doj) tuples.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    tab_name = find_tab(wb, tab_patterns)
    if not tab_name:
        wb.close()
        return []
    ws = wb[tab_name]
    col = {}
    result = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < skip_rows - 1:
            continue
        if i == skip_rows - 1:  # header row
            for idx, cell in enumerate(row):
                h = safe_str(cell).upper()
                if any(x in h for x in ('EMP ID', 'EMPLOYEE ID', 'HKID')) and 'OLD' not in h and 'PREV' not in h:
                    col['id'] = idx
                elif 'DATE OF JOINING' in h or h == 'DOJ':
                    col['doj'] = idx
                elif 'REMARKS' in h:
                    col['rem'] = idx
            # Apply fallbacks if not detected
            if 'id' not in col:
                col['id'] = id_col_fallback
            if 'doj' not in col:
                col['doj'] = doj_col_fallback
            continue

        # Data rows — get emp_id and doj
        emp_id = safe_str(row[col['id']]) if len(row) > col['id'] else ''
        if not emp_id or emp_id in ('Emp ID', 'EMP ID', 'HKID'):
            continue
        doj = row[col['doj']] if len(row) > col['doj'] else None

        # Check remarks — if labelled use that col, else scan all cols
        is_new = False
        if 'rem' in col:
            rem = safe_str(row[col['rem']]).upper() if len(row) > col['rem'] else ''
            is_new = 'NEW JOINEE' in rem
        else:
            # Unlabelled remarks (Project sheet) — scan all columns
            for cell in row:
                if cell and 'NEW JOINEE' in safe_str(cell).upper():
                    is_new = True
                    break

        if is_new:
            result.append((emp_id, doj))

    wb.close()
    return result


def get_hk_new_joinees(filepath):
    """HK new joinees come from 'New HK staff' tab — row1=title, row2=header, row3+=data."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    tab_name = find_tab(wb, ['New HK staff', 'New HK', 'New Staff'])
    if not tab_name:
        wb.close()
        return []
    ws = wb[tab_name]
    result = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # title row
        if i == 1:  # header row — detect columns dynamically
            col = detect_cols(row)
            if 'id'  not in col: col['id']  = 1  # fallback: HKID at col 1
            if 'doj' not in col: col['doj'] = 5  # fallback: DOJ at col 5
            continue
        eid = safe_str(row[col['id']]) if len(row) > col['id'] else ''
        doj = row[col['doj']] if len(row) > col['doj'] else None
        if eid and eid.upper() not in ('HKID', 'EMP ID', ''):
            result.append((eid, doj))

    wb.close()
    return result


@app.route('/download_new_joinee_excel', methods=['POST'])
def download_new_joinee_excel():
    """
    Generate formatted Excel for new joinees with DRF Care Fund data.
    Expects JSON: { "processing_month": "2026-03" }
    """

    data             = request.get_json()
    processing_month = data.get('processing_month', '')

    if not processing_month:
        return jsonify({'error': 'Processing month is required (YYYY-MM)'}), 400

    try:
        parts = processing_month.split('-')
        if len(parts) != 2 or not all(p.isdigit() for p in parts):
            raise ValueError
    except ValueError:
        return jsonify({'error': 'Invalid format — use YYYY-MM'}), 400

    # --- Collect new joinees from all sheets ---
    all_records = []
    errors      = []

    sheet_configs = [
        ('core',     ['Core Staff', 'Core'],                                                                              2, 1, 6),
        ('project',  ['Salary Request', 'Project'],                                                                       2, 1, 7),
        ('cf',       ['Active'],                                                                                          2, 1, 5),
        ('retainer', ['Retainer', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', '26'], 1, 2, 7),
        ('school',   ['KARV', 'School', 'Salary'],                                                                       2, 2, 7),
        ('college',  ['Active'],                                                                                          1, 2, 6),
    ]

    for key, tab, skip, id_fb, doj_fb in sheet_configs:
        fp = os.path.join(UPLOAD_FOLDER, f'upload_{key}.xlsx')
        if not os.path.exists(fp):
            continue
        try:
            recs = get_new_joinees_from_file(fp, tab, skip, id_fb, doj_fb)
            all_records.extend(recs)
        except Exception as e:
            errors.append(f'{key}: {e}')

    # HK — from New HK staff tab
    fp = os.path.join(UPLOAD_FOLDER, 'upload_hk.xlsx')
    if os.path.exists(fp):
        try:
            all_records.extend(get_hk_new_joinees(fp))
        except Exception as e:
            errors.append(f'hk: {e}')

    # --- Build Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'New Joinee - DRF Care Fund'

    # Styles
    HEADER_FILL  = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    ALT_FILL     = PatternFill(start_color='DEEAF1', end_color='DEEAF1', fill_type='solid')
    WHITE_FILL   = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    HEADER_FONT  = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
    DATA_FONT    = Font(size=10, name='Calibri')
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT_ALIGN   = Alignment(horizontal='left',   vertical='center')
    THIN_BORDER  = Border(
        left=Side(style='thin', color='B0C4DE'),
        right=Side(style='thin', color='B0C4DE'),
        top=Side(style='thin', color='B0C4DE'),
        bottom=Side(style='thin', color='B0C4DE'),
    )

    # Headers
    headers = [
        'Email/Employee ID',
        'Advance Category',
        'Advance Name',
        'Is Off-Cycle Payment (Yes/No)',
        'Pay Run',
        'Value Type (Total/Periodic)',
        'Total Amount/Percent/Hours/Days',
        'Currency ISO Code',
        'Frequency',
        'Start From',
        'End By',
        'Number Of Deductions',
        'Status (Open/Completed)',
        'Reason for status change',
    ]

    ws.append(headers)
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border    = THIN_BORDER

    ws.row_dimensions[1].height = 36

    # Data rows
    for row_idx, (emp_id, doj) in enumerate(all_records, start=2):
        amount   = calc_amount(doj)
        row_fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL

        row_data = [
            emp_id,           # Email/Employee ID
            'DRF Care Fund',  # Advance Category
            'DRF Care Fund',  # Advance Name
            '',               # Is Off-Cycle Payment
            '',               # Pay Run
            'Total',          # Value Type
            amount,           # Total Amount
            '',               # Currency ISO Code
            'Monthly',        # Frequency
            processing_month, # Start From
            '',               # End By
            1,                # Number Of Deductions
            '',               # Status
            '',               # Reason for status change
        ]

        ws.append(row_data)
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            cell.font      = DATA_FONT
            cell.fill      = row_fill
            cell.border    = THIN_BORDER
            cell.alignment = CENTER_ALIGN if col_idx != 1 else LEFT_ALIGN

        ws.row_dimensions[row_idx].height = 20

    # Column widths
    col_widths = [22, 18, 18, 28, 10, 26, 32, 18, 12, 12, 10, 22, 24, 24]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Summary row at bottom
    if all_records:
        ws.append([])  # blank row
        summary_row = ws.max_row + 1
        ws.cell(row=summary_row, column=1, value=f'Total New Joinees: {len(all_records)}')
        ws.cell(row=summary_row, column=1).font = Font(bold=True, size=10, name='Calibri', color='1F4E79')

    # Save
    safe_month   = processing_month.replace('-', '')
    output_path  = os.path.join(UPLOAD_FOLDER, f'NewJoinee_{safe_month}.xlsx')
    wb.save(output_path)
    wb.close()

    return send_file(
        output_path,
        as_attachment=True,
        download_name=f'NewJoinee_{safe_month}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


if __name__ == '__main__':
    from waitress import serve
    print('Serving on http://0.0.0.0:5000')
    serve(app, host='0.0.0.0', port=5000, threads=4)